#!/usr/bin/env python3
"""
API REST para el dashboard de gastos.
Sirve los datos desde la misma DB que usa el bot de Telegram.
"""

import sqlite3
import os
import io
from datetime import date
from flask import Flask, jsonify, request, send_file, send_from_directory
from flask_cors import CORS

app = Flask(__name__, static_folder=".")
CORS(app)

DB_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "gastos.db")

# ─── Utils ─────────────────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

# ─── Rutas de la API ───────────────────────────────────────────────────────────
@app.route("/")
def index():
    return send_from_directory(".", "index.html")

@app.route("/api/gastos")
def gastos():
    mes = request.args.get("mes") or date.today().strftime("%Y-%m")
    prefix = mes + "%"
    
    conn = get_db()
    cur = conn.cursor()
    
    # Gastos individuales
    cur.execute("""
        SELECT fecha, categoria, descripcion, monto
        FROM gastos
        WHERE fecha LIKE ?
        ORDER BY fecha DESC, created_at DESC
    """, (prefix,))
    rows = [dict(r) for r in cur.fetchall()]
    
    # Totales por categoría
    cur.execute("""
        SELECT categoria, SUM(monto) as total
        FROM gastos
        WHERE fecha LIKE ?
        GROUP BY categoria
        ORDER BY total DESC
    """, (prefix,))
    por_cat = {r["categoria"]: r["total"] for r in cur.fetchall()}
    
    total = sum(r["monto"] for r in rows)
    
    conn.close()
    
    return jsonify({
        "mes": mes,
        "gastos": rows,
        "por_categoria": por_cat,
        "total": total
    })

@app.route("/api/exportar")
def exportar():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter

    mes = request.args.get("mes") or date.today().strftime("%Y-%m")
    prefix = mes + "%"
    
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT fecha, categoria, descripcion, monto
        FROM gastos WHERE fecha LIKE ?
        ORDER BY fecha DESC
    """, (prefix,))
    gastos = [dict(r) for r in cur.fetchall()]
    
    cur.execute("""
        SELECT categoria, SUM(monto) as total, COUNT(*) as qty
        FROM gastos WHERE fecha LIKE ?
        GROUP BY categoria ORDER BY total DESC
    """, (prefix,))
    resumen = [dict(r) for r in cur.fetchall()]
    conn.close()
    
    wb = openpyxl.Workbook()
    
    # ── Estilos ──
    h_fill  = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
    h_font  = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
    alt_fill = PatternFill(start_color="EDF2F7", end_color="EDF2F7", fill_type="solid")
    total_fill = PatternFill(start_color="BEE3F8", end_color="BEE3F8", fill_type="solid")
    thin = Side(style="thin", color="CBD5E0")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    def header_cell(ws, row, col, val, w=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font = h_font; c.fill = h_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = brd
        if w: ws.column_dimensions[get_column_letter(col)].width = w
        return c
    
    def data_cell(ws, row, col, val, fmt=None, bold=False, fill=None):
        c = ws.cell(row=row, column=col, value=val)
        c.border = brd
        c.alignment = Alignment(vertical="center")
        if fmt: c.number_format = fmt
        if bold: c.font = Font(bold=True, name="Calibri")
        if fill: c.fill = fill
        return c
    
    # ── Hoja 1: Detalle ──
    ws1 = wb.active
    ws1.title = f"Gastos {mes}"
    ws1.row_dimensions[1].height = 28
    
    for col, (h, w) in enumerate(zip(
        ["Fecha", "Categoría", "Descripción", "Monto ($)"],
        [14, 22, 38, 16]
    ), 1):
        header_cell(ws1, 1, col, h, w)
    
    total_gral = 0
    for i, g in enumerate(gastos, 2):
        fill = alt_fill if i % 2 == 0 else None
        data_cell(ws1, i, 1, g["fecha"], fill=fill)
        data_cell(ws1, i, 2, g["categoria"], fill=fill)
        data_cell(ws1, i, 3, g["descripcion"], fill=fill)
        c = data_cell(ws1, i, 4, g["monto"], fmt='#,##0.00', fill=fill)
        c.alignment = Alignment(horizontal="right", vertical="center")
        total_gral += g["monto"]
    
    last = len(gastos) + 2
    data_cell(ws1, last, 3, "TOTAL", bold=True, fill=total_fill)
    c = data_cell(ws1, last, 4, total_gral, fmt='#,##0.00', bold=True, fill=total_fill)
    c.alignment = Alignment(horizontal="right", vertical="center")
    
    # ── Hoja 2: Resumen por categoría ──
    ws2 = wb.create_sheet("Resumen categorías")
    ws2.row_dimensions[1].height = 28
    
    for col, (h, w) in enumerate(zip(
        ["Categoría", "Total ($)", "Nro. gastos", "% del total"],
        [24, 16, 14, 14]
    ), 1):
        header_cell(ws2, 1, col, h, w)
    
    total_cat = sum(r["total"] for r in resumen)
    for i, r in enumerate(resumen, 2):
        fill = alt_fill if i % 2 == 0 else None
        data_cell(ws2, i, 1, r["categoria"], fill=fill)
        c = data_cell(ws2, i, 2, r["total"], fmt='#,##0.00', fill=fill)
        c.alignment = Alignment(horizontal="right", vertical="center")
        data_cell(ws2, i, 3, r["qty"], fill=fill)
        pct = r["total"] / total_cat * 100 if total_cat else 0
        c = data_cell(ws2, i, 4, pct/100, fmt='0.0%', fill=fill)
        c.alignment = Alignment(horizontal="right", vertical="center")
    
    # Gráfico de torta
    pie = openpyxl.chart.PieChart()
    pie.title = f"Distribución de gastos – {mes}"
    pie.style = 10
    
    labels = Reference(ws2, min_col=1, min_row=2, max_row=len(resumen)+1)
    data_ref = Reference(ws2, min_col=2, min_row=1, max_row=len(resumen)+1)
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(labels)
    pie.width = 18; pie.height = 14
    ws2.add_chart(pie, "F2")
    
    # Enviar archivo
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    nombre = f"gastos_{mes.replace('-','')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=nombre
    )

@app.route("/api/stats")
def stats():
    """Estadísticas generales para todos los meses."""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT strftime('%Y-%m', fecha) as mes, SUM(monto) as total
        FROM gastos GROUP BY mes ORDER BY mes
    """)
    por_mes = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify({"por_mes": por_mes})

if __name__ == "__main__":
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    print("🚀 Servidor iniciado en http://localhost:5000")
    app.run(debug=True, port=5000)
