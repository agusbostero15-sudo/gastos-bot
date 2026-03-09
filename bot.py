#!/usr/bin/env python3
"""
Bot de Telegram para registro de gastos personales.
Guarda gastos en SQLite y permite exportar a Excel.
"""

import logging
import re
import sqlite3
import os
from datetime import datetime, date
from telegram import Update, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    filters, ContextTypes, ConversationHandler
)

# ─── Config ───────────────────────────────────────────────────────────────────
TOKEN = os.getenv("TELEGRAM_TOKEN", "8706111305:AAE5G3aSTZo1jD5DBDF-p5ROAWtUsgWEUpY")
DB_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "gastos.db")

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Categorías disponibles
CATEGORIAS = [
    "🍔 Comida", "🚗 Transporte", "🏠 Hogar", "💊 Salud",
    "🎮 Ocio", "👔 Ropa", "📚 Educación", "💡 Servicios",
    "🛒 Supermercado", "🍺 Salidas", "✈️ Viajes", "📦 Otros"
]

ESPERANDO_CATEGORIA = 1
ESPERANDO_DESCRIPCION = 2

# ─── Base de datos ─────────────────────────────────────────────────────────────
def init_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS gastos (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id     INTEGER NOT NULL,
            username    TEXT,
            monto       REAL NOT NULL,
            categoria   TEXT NOT NULL,
            descripcion TEXT,
            fecha       TEXT NOT NULL,
            created_at  TEXT NOT NULL
        )
    """)
    conn.commit()
    conn.close()

def guardar_gasto(user_id, username, monto, categoria, descripcion, fecha=None):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    hoy = fecha or date.today().isoformat()
    cur.execute("""
        INSERT INTO gastos (user_id, username, monto, categoria, descripcion, fecha, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (user_id, username, monto, categoria, descripcion, hoy, datetime.now().isoformat()))
    conn.commit()
    conn.close()

def obtener_resumen_mes(user_id, mes=None, anio=None):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    hoy = date.today()
    mes = mes or hoy.month
    anio = anio or hoy.year
    prefix = f"{anio}-{mes:02d}"
    cur.execute("""
        SELECT categoria, SUM(monto), COUNT(*)
        FROM gastos
        WHERE user_id = ? AND fecha LIKE ?
        GROUP BY categoria
        ORDER BY SUM(monto) DESC
    """, (user_id, f"{prefix}%"))
    rows = cur.fetchall()
    total = sum(r[1] for r in rows)
    conn.close()
    return rows, total

def obtener_ultimos_gastos(user_id, limite=5):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("""
        SELECT monto, categoria, descripcion, fecha
        FROM gastos
        WHERE user_id = ?
        ORDER BY created_at DESC
        LIMIT ?
    """, (user_id, limite))
    rows = cur.fetchall()
    conn.close()
    return rows

def eliminar_ultimo_gasto(user_id):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("""
        DELETE FROM gastos WHERE id = (
            SELECT id FROM gastos WHERE user_id = ?
            ORDER BY created_at DESC LIMIT 1
        )
    """, (user_id,))
    affected = cur.rowcount
    conn.commit()
    conn.close()
    return affected > 0

# ─── Parseo de mensajes ────────────────────────────────────────────────────────
def parsear_gasto(texto):
    """
    Intenta parsear un mensaje como gasto.
    Formatos aceptados:
      - "500 comida almuerzo"
      - "comida 500 almuerzo"  
      - "gasté 1200 en transporte"
      - "supermercado $350.50"
    """
    texto = texto.lower().strip()
    
    # Buscar monto (número con o sin $, punto o coma como decimal)
    monto_match = re.search(r'\$?\s*(\d+(?:[.,]\d{1,2})?)', texto)
    if not monto_match:
        return None, None, None
    
    monto_str = monto_match.group(1).replace(".", "").replace(",", ".")
    monto = float(monto_str)
    
    # Detectar categoría por palabras clave
    categoria_map = {
        "🍔 Comida":       ["comida", "almuerzo", "cena", "desayuno", "restaurant", "pizza", "hamburgues", "delivery", "pedidos"],
        "🚗 Transporte":   ["transporte", "colectivo", "uber", "taxi", "nafta", "combustible", "tren", "subte", "bus"],
        "🏠 Hogar":        ["hogar", "alquiler", "casa", "expensas", "muebles", "reparacion"],
        "💊 Salud":        ["salud", "medico", "farmacia", "medicina", "doctor", "clinica", "turno"],
        "🎮 Ocio":         ["ocio", "juego", "netflix", "spotify", "cine", "entretenimiento", "streaming"],
        "👔 Ropa":         ["ropa", "zapatillas", "calzado", "indumentaria", "zapatos"],
        "📚 Educación":    ["educacion", "curso", "libro", "estudio", "universidad", "colegio"],
        "💡 Servicios":    ["servicio", "luz", "agua", "gas", "internet", "telefono", "celular"],
        "🛒 Supermercado": ["super", "supermercado", "verduleria", "carniceria", "almacen", "mercado"],
        "🍺 Salidas":      ["salida", "bar", "boliche", "cerveza", "tragos", "fernet", "birra"],
        "✈️ Viajes":       ["viaje", "vuelo", "hotel", "turismo", "vacaciones"],
        "📦 Otros":        ["otro", "varios", "misc"],
    }
    
    categoria_detectada = None
    for cat, palabras in categoria_map.items():
        for p in palabras:
            if p in texto:
                categoria_detectada = cat
                break
        if categoria_detectada:
            break
    
    # Descripción: el texto sin el monto
    descripcion = re.sub(r'\$?\s*\d+(?:[.,]\d{1,2})?', '', texto).strip()
    descripcion = re.sub(r'\b(gaste|gasté|en|de|por|pague|pagué)\b', '', descripcion).strip()
    descripcion = descripcion.capitalize() or "Sin descripción"
    
    return monto, categoria_detectada, descripcion

# ─── Teclado ───────────────────────────────────────────────────────────────────
def teclado_categorias():
    botones = []
    row = []
    for i, cat in enumerate(CATEGORIAS):
        row.append(KeyboardButton(cat))
        if len(row) == 2:
            botones.append(row)
            row = []
    if row:
        botones.append(row)
    botones.append([KeyboardButton("❌ Cancelar")])
    return ReplyKeyboardMarkup(botones, resize_keyboard=True, one_time_keyboard=True)

def teclado_principal():
    botones = [
        [KeyboardButton("📊 Resumen del mes"), KeyboardButton("📋 Últimos gastos")],
        [KeyboardButton("📥 Exportar Excel"), KeyboardButton("🗑️ Deshacer último")],
        [KeyboardButton("❓ Ayuda")]
    ]
    return ReplyKeyboardMarkup(botones, resize_keyboard=True)

# ─── Handlers ─────────────────────────────────────────────────────────────────
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    nombre = update.effective_user.first_name
    await update.message.reply_text(
        f"👋 ¡Hola *{nombre}*! Soy tu asistente de gastos.\n\n"
        "Podés registrar un gasto de varias formas:\n"
        "• Simplemente escribí: *500 comida almuerzo*\n"
        "• O: *gasté 1200 en transporte*\n"
        "• O usá el comando /nuevo para elegir categoría manualmente\n\n"
        "📌 También podés usar los botones de abajo para ver resúmenes y exportar.",
        parse_mode="Markdown",
        reply_markup=teclado_principal()
    )

async def ayuda(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "📖 *Cómo usar el bot:*\n\n"
        "*Registrar un gasto rápido:*\n"
        "  `500 comida almuerzo`\n"
        "  `1200 transporte uber`\n"
        "  `$350.50 supermercado`\n"
        "  `gasté 800 en salidas`\n\n"
        "*Comandos disponibles:*\n"
        "  /nuevo → Registro manual con categoría\n"
        "  /resumen → Resumen del mes actual\n"
        "  /ultimos → Últimos 5 gastos\n"
        "  /exportar → Descargar Excel\n"
        "  /deshacer → Eliminar último gasto\n"
        "  /start → Volver al inicio\n\n"
        "*Categorías disponibles:*\n"
        + "  " + " | ".join(CATEGORIAS),
        parse_mode="Markdown",
        reply_markup=teclado_principal()
    )

async def nuevo_gasto_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Inicia el flujo de registro manual."""
    context.user_data.clear()
    await update.message.reply_text(
        "💰 *¿Cuánto gastaste?*\nEscribí el monto (ej: `500` o `1350.50`)",
        parse_mode="Markdown"
    )
    return ESPERANDO_CATEGORIA

async def recibir_monto(update: Update, context: ContextTypes.DEFAULT_TYPE):
    texto = update.message.text.replace("$", "").replace(",", ".").strip()
    try:
        monto = float(texto)
        context.user_data["monto"] = monto
        await update.message.reply_text(
            f"✅ Monto: *${monto:,.2f}*\n\n📂 Ahora elegí la categoría:",
            parse_mode="Markdown",
            reply_markup=teclado_categorias()
        )
        return ESPERANDO_DESCRIPCION
    except ValueError:
        await update.message.reply_text("❌ Eso no parece un número. Escribí solo el monto, ej: `500`")
        return ESPERANDO_CATEGORIA

async def recibir_categoria(update: Update, context: ContextTypes.DEFAULT_TYPE):
    texto = update.message.text
    if texto == "❌ Cancelar":
        await update.message.reply_text("❌ Cancelado.", reply_markup=teclado_principal())
        return ConversationHandler.END
    
    if texto not in CATEGORIAS:
        await update.message.reply_text("Por favor elegí una categoría del teclado.")
        return ESPERANDO_DESCRIPCION
    
    context.user_data["categoria"] = texto
    await update.message.reply_text(
        f"📝 Categoría: *{texto}*\n\nAgregá una descripción corta (o escribí `-` para omitir):",
        parse_mode="Markdown"
    )
    return ConversationHandler.END + 1  # paso extra

# Usamos un handler directo para el fin del flujo
async def recibir_descripcion(update: Update, context: ContextTypes.DEFAULT_TYPE):
    descripcion = update.message.text
    if descripcion == "-":
        descripcion = "Sin descripción"
    
    monto = context.user_data.get("monto")
    categoria = context.user_data.get("categoria")
    user = update.effective_user
    
    guardar_gasto(user.id, user.username or user.first_name, monto, categoria, descripcion)
    
    await update.message.reply_text(
        f"✅ *Gasto registrado*\n\n"
        f"💰 Monto: `${monto:,.2f}`\n"
        f"📂 Categoría: {categoria}\n"
        f"📝 Descripción: {descripcion}\n"
        f"📅 Fecha: {date.today().strftime('%d/%m/%Y')}",
        parse_mode="Markdown",
        reply_markup=teclado_principal()
    )
    context.user_data.clear()
    return ConversationHandler.END

async def cancelar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text("❌ Cancelado.", reply_markup=teclado_principal())
    return ConversationHandler.END

async def resumen(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    rows, total = obtener_resumen_mes(user_id)
    hoy = date.today()
    
    if not rows:
        await update.message.reply_text(
            f"📊 No hay gastos registrados para *{hoy.strftime('%B %Y')}*.",
            parse_mode="Markdown"
        )
        return
    
    texto = f"📊 *Resumen de {hoy.strftime('%B %Y')}*\n\n"
    for cat, suma, count in rows:
        porcentaje = (suma / total * 100) if total > 0 else 0
        barra = "█" * int(porcentaje / 10) + "░" * (10 - int(porcentaje / 10))
        texto += f"{cat}\n`{barra}` {porcentaje:.0f}%\n💵 ${suma:,.2f} ({count} gastos)\n\n"
    
    texto += f"━━━━━━━━━━━━━━\n💰 *Total: ${total:,.2f}*"
    
    await update.message.reply_text(texto, parse_mode="Markdown", reply_markup=teclado_principal())

async def ultimos_gastos(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    gastos = obtener_ultimos_gastos(user_id)
    
    if not gastos:
        await update.message.reply_text("📋 No tenés gastos registrados aún.")
        return
    
    texto = "📋 *Últimos 5 gastos:*\n\n"
    for monto, cat, desc, fecha in gastos:
        fecha_fmt = datetime.strptime(fecha, "%Y-%m-%d").strftime("%d/%m")
        texto += f"• `{fecha_fmt}` {cat} — *${monto:,.2f}*\n  _{desc}_\n\n"
    
    await update.message.reply_text(texto, parse_mode="Markdown", reply_markup=teclado_principal())

async def exportar_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter
    import io

    user_id = update.effective_user.id
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    
    # Todos los gastos del usuario
    cur.execute("""
        SELECT fecha, categoria, descripcion, monto
        FROM gastos WHERE user_id = ?
        ORDER BY fecha DESC
    """, (user_id,))
    gastos = cur.fetchall()
    conn.close()
    
    if not gastos:
        await update.message.reply_text("❌ No tenés gastos para exportar.")
        return
    
    wb = openpyxl.Workbook()
    
    # ── Hoja 1: Todos los gastos ──
    ws = wb.active
    ws.title = "Todos los gastos"
    
    # Estilos
    header_fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    alt_fill = PatternFill(start_color="EDF2F7", end_color="EDF2F7", fill_type="solid")
    thin = Side(style="thin", color="CBD5E0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    headers = ["Fecha", "Categoría", "Descripción", "Monto ($)"]
    col_widths = [15, 22, 35, 15]
    
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 25
    
    total = 0
    for i, (fecha, cat, desc, monto) in enumerate(gastos, 2):
        fecha_fmt = datetime.strptime(fecha, "%Y-%m-%d").strftime("%d/%m/%Y")
        row_data = [fecha_fmt, cat, desc, monto]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if i % 2 == 0:
                cell.fill = alt_fill
            if col == 4:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
        total += monto
    
    # Fila total
    last_row = len(gastos) + 2
    ws.cell(row=last_row, column=3, value="TOTAL").font = Font(bold=True)
    cell_total = ws.cell(row=last_row, column=4, value=total)
    cell_total.font = Font(bold=True, color="1A202C")
    cell_total.number_format = '#,##0.00'
    cell_total.fill = PatternFill(start_color="BEE3F8", end_color="BEE3F8", fill_type="solid")
    
    # ── Hoja 2: Resumen por mes ──
    ws2 = wb.create_sheet("Resumen mensual")
    cur2 = sqlite3.connect(DB_PATH).cursor()
    cur2.execute("""
        SELECT strftime('%Y-%m', fecha) as mes, categoria, SUM(monto)
        FROM gastos WHERE user_id = ?
        GROUP BY mes, categoria
        ORDER BY mes DESC, SUM(monto) DESC
    """, (user_id,))
    resumen_data = cur2.fetchall()
    
    ws2.cell(row=1, column=1, value="Mes").font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=2, value="Categoría").font = header_font
    ws2.cell(row=1, column=2).fill = header_fill
    ws2.cell(row=1, column=3, value="Total ($)").font = header_font
    ws2.cell(row=1, column=3).fill = header_fill
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 22
    ws2.column_dimensions['C'].width = 15
    
    for i, (mes, cat, suma) in enumerate(resumen_data, 2):
        ws2.cell(row=i, column=1, value=mes)
        ws2.cell(row=i, column=2, value=cat)
        cell = ws2.cell(row=i, column=3, value=suma)
        cell.number_format = '#,##0.00'
        if i % 2 == 0:
            for c in range(1, 4):
                ws2.cell(row=i, column=c).fill = alt_fill
    
    # Guardar y enviar
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    nombre_archivo = f"gastos_{date.today().strftime('%Y%m')}.xlsx"
    await update.message.reply_document(
        document=buffer,
        filename=nombre_archivo,
        caption=f"📊 *Tu reporte de gastos*\n{len(gastos)} registros exportados.",
        parse_mode="Markdown",
        reply_markup=teclado_principal()
    )

async def deshacer(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if eliminar_ultimo_gasto(user_id):
        await update.message.reply_text(
            "🗑️ Último gasto eliminado.",
            reply_markup=teclado_principal()
        )
    else:
        await update.message.reply_text("❌ No hay gastos para eliminar.")

async def mensaje_libre(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Maneja mensajes de texto libre intentando parsearlos como gastos."""
    texto = update.message.text
    
    # Botones del teclado principal
    if texto == "📊 Resumen del mes":
        return await resumen(update, context)
    if texto == "📋 Últimos gastos":
        return await ultimos_gastos(update, context)
    if texto == "📥 Exportar Excel":
        return await exportar_excel(update, context)
    if texto == "🗑️ Deshacer último":
        return await deshacer(update, context)
    if texto == "❓ Ayuda":
        return await ayuda(update, context)
    
    # Intentar parsear como gasto
    monto, categoria, descripcion = parsear_gasto(texto)
    
    if monto and monto > 0:
        user = update.effective_user
        
        if categoria:
            # Gasto con categoría detectada automáticamente
            guardar_gasto(user.id, user.username or user.first_name, monto, categoria, descripcion)
            await update.message.reply_text(
                f"✅ *Gasto registrado automáticamente*\n\n"
                f"💰 ${monto:,.2f} — {categoria}\n"
                f"📝 {descripcion}\n\n"
                f"_¿Incorrecto? Usá /deshacer y registralo con /nuevo_",
                parse_mode="Markdown",
                reply_markup=teclado_principal()
            )
        else:
            # Monto detectado pero sin categoría → pedir categoría
            context.user_data["monto"] = monto
            context.user_data["descripcion"] = descripcion
            context.user_data["esperando_cat"] = True
            await update.message.reply_text(
                f"💰 Monto detectado: *${monto:,.2f}*\n\n¿En qué categoría lo pongo?",
                parse_mode="Markdown",
                reply_markup=teclado_categorias()
            )
    elif context.user_data.get("esperando_cat"):
        # Recibiendo categoría para gasto anterior
        if texto in CATEGORIAS:
            monto = context.user_data["monto"]
            desc = context.user_data.get("descripcion", "Sin descripción")
            user = update.effective_user
            guardar_gasto(user.id, user.username or user.first_name, monto, texto, desc)
            context.user_data.clear()
            await update.message.reply_text(
                f"✅ *Gasto registrado*\n\n"
                f"💰 ${monto:,.2f} — {texto}\n"
                f"📝 {desc}",
                parse_mode="Markdown",
                reply_markup=teclado_principal()
            )
        elif texto == "❌ Cancelar":
            context.user_data.clear()
            await update.message.reply_text("❌ Cancelado.", reply_markup=teclado_principal())
    else:
        await update.message.reply_text(
            "🤔 No entendí ese mensaje.\n\n"
            "Para registrar un gasto escribí algo como:\n"
            "`500 comida almuerzo`\n"
            "`1200 transporte uber`\n\n"
            "O usá /nuevo para el modo manual.",
            parse_mode="Markdown",
            reply_markup=teclado_principal()
        )

# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    init_db()
    
    app = Application.builder().token(TOKEN).build()
    
    # ConversationHandler para /nuevo
    conv = ConversationHandler(
        entry_points=[CommandHandler("nuevo", nuevo_gasto_start)],
        states={
            ESPERANDO_CATEGORIA: [MessageHandler(filters.TEXT & ~filters.COMMAND, recibir_monto)],
            ESPERANDO_DESCRIPCION: [MessageHandler(filters.TEXT & ~filters.COMMAND, recibir_categoria)],
            ConversationHandler.END + 1: [MessageHandler(filters.TEXT & ~filters.COMMAND, recibir_descripcion)],
        },
        fallbacks=[CommandHandler("cancelar", cancelar)],
    )
    
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("ayuda", ayuda))
    app.add_handler(CommandHandler("resumen", resumen))
    app.add_handler(CommandHandler("ultimos", ultimos_gastos))
    app.add_handler(CommandHandler("exportar", exportar_excel))
    app.add_handler(CommandHandler("deshacer", deshacer))
    app.add_handler(conv)
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, mensaje_libre))
    
    logger.info("🤖 Bot iniciado. Esperando mensajes...")
    app.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == "__main__":
    main()
